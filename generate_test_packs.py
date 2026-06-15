#!/usr/bin/env python3
"""
Generate the per-tester file packs for the platform test program.

Each tester gets a folder under test-packs/ with exactly the files their
checklist needs, plus a README. Trish's document filenames use REAL demo
asset FHA case numbers (from demo/data.js) so they classify to actual
assets in the HVLS-2026-DEMO data room. Run: python generate_test_packs.py
"""
import os, struct, random, datetime

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "test-packs")

# Real HVLS-2026-DEMO assets (state, FHA case number) pulled from demo/data.js.
ASSETS = [
    ("GA", "215-8969045"),
    ("FL", "869-7085471"),
    ("CA", "585-6141309"),
]

# ---------------------------------------------------------------- minimal PDF
def _ascii(s):
    return (str(s).replace("—", "-").replace("–", "-")
            .replace("’", "'").replace("‘", "'")
            .replace("“", '"').replace("”", '"'))

def _esc(s):
    return _ascii(s).replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")

def minimal_pdf(title, lines):
    content = "BT /F1 16 Tf 50 742 Td (%s) Tj ET\n" % _esc(title)
    y = 712
    for ln in lines:
        content += "BT /F1 11 Tf 50 %d Td (%s) Tj ET\n" % (y, _esc(ln))
        y -= 16
    objs = [
        "<< /Type /Catalog /Pages 2 0 R >>",
        "<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        "<< /Length %d >>\nstream\n%sendstream" % (len(content), content),
        "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out = "%PDF-1.4\n"
    offsets = []
    for i, o in enumerate(objs, start=1):
        offsets.append(len(out.encode("latin-1")))
        out += "%d 0 obj\n%s\nendobj\n" % (i, o)
    xref = len(out.encode("latin-1"))
    out += "xref\n0 %d\n0000000000 65535 f \n" % (len(objs) + 1)
    for off in offsets:
        out += "%010d 00000 n \n" % off
    out += "trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF" % (len(objs) + 1, xref)
    return out.encode("latin-1")

def write_pdf(folder, name, title, lines):
    with open(os.path.join(folder, name), "wb") as f:
        f.write(minimal_pdf(title, lines))

def ensure(*parts):
    p = os.path.join(*parts)
    os.makedirs(p, exist_ok=True)
    return p

STAMP = "Sample document for platform testing. Not a real loan file."

# ---------------------------------------------------------------- Trish pack
def trish():
    d = ensure(OUT, "trish-dataroom")
    made = []
    def doc(name, kind):
        write_pdf(d, name, kind, [
            "Sale: HUD HECM Vacant Loan Sale (HVLS-2026-DEMO)",
            "File: " + name,
            "Document type: " + kind,
            "", STAMP,
        ])
        made.append(name)

    # Collateral + Due Diligence for three real assets (filenames carry the real FHA#)
    for st, fha in ASSETS:
        doc("%s_%s_Note.pdf" % (st, fha), "Promissory Note")
        doc("%s_%s_Mortgage.pdf" % (st, fha), "Mortgage / Deed of Trust")
        doc("%s_%s_BPO.pdf" % (st, fha), "Broker Price Opinion")
        doc("%s_%s_Title_Search.pdf" % (st, fha), "Title Search")
    # camelCase collateral (tests the camelCase split -> Collateral)
    st, fha = ASSETS[0]
    doc("%s_%s_AssignmentOfMortgage.pdf" % (st, fha), "Assignment of Mortgage")
    doc("%s_%s_Servicing_Comments.pdf" % (ASSETS[1][0], ASSETS[1][1]), "Servicing Comments")
    doc("%s_%s_Occupancy_Inspection.pdf" % (ASSETS[2][0], ASSETS[2][1]), "Occupancy / Inspection")

    # Sale-level documents (route to folders by name, no FHA#)
    doc("HVLS-2026-DEMO-Bidder-Information-Package.pdf", "Bidder Information Package")
    doc("HVLS-2026-DEMO-ALD-Loan-Tape.pdf", "Loan Tape (ALD)")
    doc("HVLS-2026-DEMO-Sale-and-Bid-Day-Procedures.pdf", "Procedures")
    doc("HVLS-2026-DEMO-BAUF-BTAF-Deposit-Forms.pdf", "Forms & Agreements")

    # Should land in the REVIEW queue (no asset match, no folder keyword)
    doc("Scanned-Document-4471.pdf", "Unlabeled scan")
    doc("misc-loan-notes.pdf", "Miscellaneous")

    # Awkward names for TD11 (real FHA#, odd delimiters) — should still match
    doc("GA-215-8969045-Note.pdf", "Promissory Note (dash-delimited)")
    doc("585-6141309_CA_BPO.pdf", "BPO (reordered name)")

    readme = """TRISH — DATA ROOM OPERATIONS · YOUR TEST FILES
================================================================

These are sample documents to drag into the Admin > Data Room (the staff
intake screen) for the HVLS-2026-DEMO sale. The platform reads the FILE NAME
to decide where each one belongs, so the names here are built to the real
{STATE}_{FHA#}_{DOCTYPE} convention and use real case numbers from that sale.

HOW TO USE (cases TD1-TD5, TD11)
  1. Open Admin > Data Room and pick the HVLS sale.
  2. Drag this whole folder's PDFs onto the drop zone (or browse to them).
  3. Watch where each one is auto-filed, then compare to "WHAT TO EXPECT".

WHAT TO EXPECT
  Collateral (under the matching asset):
     *_Note.pdf, *_Mortgage.pdf, *_AssignmentOfMortgage.pdf
  Due Diligence (under the matching asset):
     *_BPO.pdf, *_Title_Search.pdf, *_Servicing_Comments.pdf, *_Occupancy_Inspection.pdf
  Sale folders (no asset, filed by name):
     ...Bidder-Information-Package.pdf  -> Bidder Information Package
     ...ALD-Loan-Tape.pdf               -> Loan Tape
     ...Procedures.pdf                  -> Procedures
     ...BAUF-BTAF-Deposit-Forms.pdf     -> Forms & Agreements
  Review queue (could not be matched, needs your decision):
     Scanned-Document-4471.pdf, misc-loan-notes.pdf
  Awkward but still matched (TD11):
     GA-215-8969045-Note.pdf  (all dashes) -> Collateral on that asset
     585-6141309_CA_BPO.pdf   (reordered)  -> Due Diligence on that asset

Mark Fail on anything that files in the wrong place, and note the exact file
name. For TD11, rename copies of these to your own awkward variants and see
what the classifier does.

The Healthcare data-room check (TD9) and the Ginnie Mae check (TD10) do NOT
need uploads — those documents are already in those rooms; you just open and
read them.
"""
    open(os.path.join(d, "README.txt"), "w", encoding="utf-8").write(readme)
    return len(made)

# ---------------------------------------------------------------- Jelani tape
def jelani():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    d = ensure(OUT, "jelani-admin")
    wb = Workbook(); ws = wb.active; ws.title = "Loan Tape"
    headers = ["Loan ID", "FHA case number", "Current UPB", "Unpaid loan balance (ULB)",
               "BPO value", "ETD-adjusted BPO", "Max claim amount", "Original principal balance",
               "Interest rate", "Property address", "Property city", "Property state",
               "Property ZIP", "Occupancy status", "Property condition",
               "Property / project name", "Asset class", "Units", "Pool (grouping)"]
    ws.append(headers)
    hf = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1E1F6B")
    for c in ws[1]:
        c.font = hf; c.fill = fill; c.alignment = Alignment(horizontal="center")
    rng = random.Random(20260615)
    GEO = {"FL": "Jacksonville", "TX": "Houston", "CA": "Fresno", "GA": "Atlanta",
           "OH": "Toledo", "PA": "Erie", "NC": "Raleigh", "AZ": "Mesa", "MI": "Flint", "IL": "Peoria"}
    states = list(GEO.keys())
    streets = ["Magnolia Blvd", "Live Oak Dr", "Cedar St", "Maple Ave", "Sycamore Ln",
               "Willow Creek Rd", "Hickory Ln", "Dogwood Dr", "Birch Way", "Aspen Ct"]
    N = 140
    for i in range(N):
        st = rng.choice(states)
        upb = round((90000 + rng.random() * 320000) / 1000) * 1000
        ulb = round(upb * (1.03 + rng.random() * 0.12) / 1000) * 1000
        bpo = round(upb / (0.32 + rng.random() * 0.38) / 1000) * 1000
        etd = round(bpo * (0.85 + rng.random() * 0.40) / 1000) * 1000
        mca = round(upb * (1.0 + rng.random() * 0.16) / 1000) * 1000
        opb = round(upb * (0.88 + rng.random() * 0.16) / 1000) * 1000
        rate = round(3 + rng.random() * 4, 3)
        addr = "%d %s" % (100 + int(rng.random() * 8900), rng.choice(streets))
        zipc = "%05d" % (30000 + int(rng.random() * 60000))
        occ = "VACANT" if rng.random() < 0.66 else "OCCUPIED"
        cond = rng.choice(["GOOD", "GOOD", "FAIR", "FAIR", "POOR"])
        pool = "Pool %d" % (1 + (i % 2))
        ws.append(["L-%05d" % (10001 + i), "%03d-%07d" % (100 + int(rng.random()*800), 1000000 + int(rng.random()*8999999)),
                   upb, ulb, bpo, etd, mca, opb, rate, addr, GEO[st], st, zipc, occ, cond, "", "", "", pool])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 17
    ws.freeze_panes = "A2"
    wb.save(os.path.join(d, "SAMPLE-HVLS-Tape.xlsx"))

    readme = """JELANI — ADMINISTRATOR · YOUR TEST FILES
================================================================

SAMPLE-HVLS-Tape.xlsx
  A realistic HECM loan tape (140 loans, 19 columns) in the exact format the
  Sale Setup tool expects. Use it to test the real tape-upload path instead of
  the built-in "Use sample data".

HOW TO USE
  Demo (case JA3 variant): Admin > Sale Setup > pick HVLS > Upload tape >
     choose this file. Every column should auto-map; ~140 loans parse; pools
     and aggregates compute; ULB sits above UPB.
  Live (case JA15, REAL): sign in, run the same upload, Create the sale, then
     delete it when done so production stays clean.

The "Download tape template" button in Sale Setup produces the same column
layout if you want a blank to hand a real seller later.
"""
    open(os.path.join(d, "README.txt"), "w", encoding="utf-8").write(readme)

# ---------------------------------------------------------------- Gilda worksheet
def gilda():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    d = ensure(OUT, "gilda-data-integrity")
    wb = Workbook()
    navy = PatternFill("solid", fgColor="1E1F6B"); hf = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    ks = wb.active; ks.title = "Known Figures"
    ks.append(["Check", "Sale", "BID %", "Basis", "Expected result"])
    for c in ks[1]:
        c.font = hf; c.fill = navy
    rows = [
        ["GD9 total", "HVLS-2026-DEMO", "55.12345", "ULB (per loan)", "Pool total = $619,588"],
        ["GD9 deposit", "HLS-2026-DEMO", "72.5", "UPB ($40,024,350)", "Deposit = $4,002,435 (10% beats floor)"],
        ["GD5 floor", "small pool", "any", "small aggregate", "Deposit = $100,000 floor"],
        ["GD6 code (HVLS)", "HVLS-2026-DEMO", "-", "-", "Completion CODE = HVLS26D742"],
        ["GD6 code (HLS)", "HLS-2026-DEMO", "-", "-", "Completion CODE = HLS26DHC84"],
        ["GD3 basis", "HVLS", "-", "-", "Basis is ULB, not BPO; label reads 'of ULB'"],
    ]
    for r in rows:
        ks.append(r)
    for w, col in zip([16, 18, 12, 22, 40], "ABCDE"):
        ks.column_dimensions[col].width = w

    cs = wb.create_sheet("Calculator")
    cs.append(["Plug in what the platform shows; the green cells compute what it SHOULD be."])
    cs["A1"].font = bold
    cs.append([])
    cs.append(["Per-loan derived bid"])
    cs["A3"].font = bold
    cs.append(["BID % (you enter)", 55.12345])
    cs.append(["Loan basis $ (you enter, e.g. ULB)", 200000])
    cs.append(["Expected BID $ (computed)", "=B4/100*B5"])
    cs.append([])
    cs.append(["Deposit on an aggregate bid"])
    cs["A8"].font = bold
    cs.append(["Aggregate bid $ (you enter)", 619588])
    cs.append(["Expected deposit (computed)", "=MAX(100000,CEILING(B9*0.1,1))"])
    cs.append(["Note: if the aggregate is below $100,000, the deposit is 50% of it instead."])
    for c in (cs["B6"], cs["B10"]):
        c.fill = PatternFill("solid", fgColor="E7F6EC"); c.font = bold
    cs.column_dimensions["A"].width = 38; cs.column_dimensions["B"].width = 20
    wb.save(os.path.join(d, "Bid-Math-Verification.xlsx"))

    readme = """GILDA — DATA INTEGRITY & EVALUATION · YOUR TEST FILES
================================================================

Bid-Math-Verification.xlsx
  Tab 1 "Known Figures": the exact numbers the platform should produce on the
     demo sales. Reproduce each one and mark Pass/Fail against it.
  Tab 2 "Calculator": type in what the platform shows (a BID %, a loan's basis
     $, or an aggregate) and the green cells show what the platform SHOULD
     compute. If the platform's number and the green cell disagree, that is a
     Fail; note both numbers.

You also use the platform's own BEM Excel (case GD7): generate it from
Admin > BEM & Awards and change a reserve cell to confirm the awards recompute
from formulas rather than being typed in.
"""
    open(os.path.join(d, "README.txt"), "w", encoding="utf-8").write(readme)

# ---------------------------------------------------------------- Team + Maurice
def team():
    d = ensure(OUT, "team-bidders")
    rng = random.Random(77)
    pcts = ["%.3f" % (46 + rng.random() * 12) for _ in range(40)]
    open(os.path.join(d, "Paste-Sample-Percentages.txt"), "w", encoding="utf-8").write("\n".join(pcts) + "\n")
    readme = """TRISH'S TEAM — COMPETING BIDDERS · YOUR TEST FILES
================================================================

Paste-Sample-Percentages.txt
  A column of 40 bid percentages for the PASTE input (case TT4). On a pool's
  bid sheet, use "Paste from spreadsheet" and paste this column; it fills the
  loans in order. (A full submission needs one percentage per loan, so on a
  large pool this fills the first 40 and the pool stays Incomplete — that is
  fine; it shows the paste mapping works. For a full bid, use "apply to all"
  or the Excel workbook.)

The Excel bid workbook (case TT3) is DOWNLOADED from the bid sheet itself
("Download bid workbook"); fill the BID % column, save, and upload it back.
There is no workbook to pre-supply — testing the download-fill-upload loop is
the point.

Spread out: different people on different browsers and phones (case TT8).
"""
    open(os.path.join(d, "README.txt"), "w", encoding="utf-8").write(readme)

def maurice():
    d = ensure(OUT, "maurice-bidder")
    readme = """MAURICE — BIDDER EXPERIENCE · YOUR TEST FILES
================================================================

You do not need any supplied files. Everything you need is in the platform:
  - You download the Excel bid workbook FROM the bid sheet (case MB8).
  - You download data-room documents FROM the Data Room (case MB5).

Your job is to go through the whole bidder journey as a sharp outside bidder
and judge whether it is clear, correct, and credible. Open your checklist in
the test console, pick "Maurice", and work top to bottom. Mark Fail on
anything that looks unfinished, reads wrong, or feels off, and say why.
"""
    open(os.path.join(d, "README.txt"), "w", encoding="utf-8").write(readme)

# ---------------------------------------------------------------- top-level
def index():
    txt = """PLATFORM TEST FILE PACKS
================================================================
Generated %s

One folder per tester. Hand each person their folder (or their zip from the
test console). Each README explains which files map to which test cases.

  jelani-admin/          loan tape for Sale Setup (real-upload test)
  trish-dataroom/        ~21 sample documents to classify in the Data Room
  gilda-data-integrity/  bid-math verification worksheet
  team-bidders/          paste-sample percentages for the paste bid input
  maurice-bidder/        no files needed (instructions only)

The master plan is platform/PLATFORM_TEST_PLAN.md. The interactive checklist
is at /demo/platform-test.html.
""" % datetime.date.today().isoformat()
    open(os.path.join(OUT, "README.md"), "w", encoding="utf-8").write(txt)

if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    n = trish(); jelani(); gilda(); team(); maurice(); index()
    print("Trish docs:", n)
    print("Wrote packs to", OUT)
